"""Excel (.xlsx) source/sink (ADR-0010, SPEC-panel §5.1/§5.3) via openpyxl.

Config per endpoint: `path`, `sheet` (name, default first sheet), optional A1
`range` (e.g. `"A1:D50"`) to scope which rows/columns are read. Mapping is
`canonical -> column`, where column is either the header name (first row of
the range) or a spreadsheet column letter (`A`, `B`, ...).
"""

from typing import Any

import openpyxl

from app.io.base import REQUIRED_PRODUCT_FIELDS, Mapping
from app.io.mapping import apply_mapping, resolve_column, reverse_apply_mapping, validate


def _load_sheet_rows(
    path: str, sheet: str | None, cell_range: str | None
) -> tuple[list[str], list[dict[str, Any]]]:
    workbook = openpyxl.load_workbook(path, data_only=True, read_only=True)
    worksheet = workbook[sheet] if sheet else workbook.worksheets[0]

    cell_rows = list(worksheet[cell_range] if cell_range else worksheet.iter_rows())
    if not cell_rows:
        return [], []

    header = [str(cell.value) if cell.value is not None else "" for cell in cell_rows[0]]
    rows: list[dict[str, Any]] = []
    for cell_row in cell_rows[1:]:
        values = [cell.value for cell in cell_row]
        if all(v is None for v in values):
            continue
        rows.append(dict(zip(header, values, strict=True)))
    return header, rows


def _mapping_to_header_names(mapping: Mapping, header: list[str]) -> Mapping:
    """Resolve each locator (header name or column letter) to its actual header cell name."""
    return {
        field: (locator if locator in header else header[resolve_column(locator, header)])
        for field, locator in mapping.items()
    }


class XlsxProductSource:
    """Reads a products sheet and a regions sheet through their `canonical -> column` mappings."""

    def __init__(
        self,
        products_path: str | None,
        products_sheet: str | None,
        products_range: str | None,
        regions_path: str | None,
        regions_sheet: str | None,
        regions_range: str | None,
        products_mapping: Mapping | None,
        regions_mapping: Mapping | None,
    ) -> None:
        self._products_path = products_path
        self._products_sheet = products_sheet
        self._products_range = products_range
        self._regions_path = regions_path
        self._regions_sheet = regions_sheet
        self._regions_range = regions_range
        self._products_mapping = products_mapping or {}
        self._regions_mapping = regions_mapping or {}

    async def read_products(self) -> list[dict[str, Any]]:
        if not self._products_path:
            return []
        header, rows = _load_sheet_rows(self._products_path, self._products_sheet, self._products_range)
        name_mapping = _mapping_to_header_names(self._products_mapping, header)
        validate(name_mapping, header, required=REQUIRED_PRODUCT_FIELDS)
        return [apply_mapping(name_mapping, row) for row in rows]

    async def read_regions(self) -> list[dict[str, Any]]:
        if not self._regions_path:
            return []
        header, rows = _load_sheet_rows(self._regions_path, self._regions_sheet, self._regions_range)
        name_mapping = _mapping_to_header_names(self._regions_mapping, header)
        validate(name_mapping, header, required=())
        return [apply_mapping(name_mapping, row) for row in rows]


class XlsxResultSink:
    """Writes canonical result rows to an xlsx sheet through the `canonical -> column` mapping."""

    def __init__(self, path: str, sheet: str | None, mapping: Mapping) -> None:
        self._path = path
        self._sheet = sheet or "Sheet1"
        self._mapping = mapping

    async def write_snapshots(self, rows: list[dict[str, Any]]) -> int:
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        assert worksheet is not None
        worksheet.title = self._sheet

        fieldnames = list(self._mapping.values())
        worksheet.append(fieldnames)
        for row in rows:
            mapped = reverse_apply_mapping(self._mapping, row)
            worksheet.append(["" if mapped[col] is None else str(mapped[col]) for col in fieldnames])

        workbook.save(self._path)
        return len(rows)
