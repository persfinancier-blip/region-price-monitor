"""app.io.xlsx_io — round-trip with a non-canonical header; Decimal prices survive as strings."""

from decimal import Decimal

import openpyxl

from app.io.xlsx_io import XlsxProductSource, XlsxResultSink


async def test_xlsx_product_source_maps_non_canonical_header_to_canonical(tmp_path) -> None:
    path = tmp_path / "products.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Товары"
    sheet.append(["Площадка", "Артикул", "Ссылка", "Название"])
    sheet.append(["wb", "12345", "https://x/1", "Товар А"])
    workbook.save(path)

    source = XlsxProductSource(
        products_path=str(path),
        products_sheet="Товары",
        products_range=None,
        regions_path=None,
        regions_sheet=None,
        regions_range=None,
        products_mapping={
            "marketplace": "Площадка",
            "sku": "Артикул",
            "url": "Ссылка",
            "name": "Название",
        },
        regions_mapping=None,
    )

    rows = await source.read_products()

    assert rows == [
        {"marketplace": "wb", "sku": "12345", "url": "https://x/1", "name": "Товар А"},
    ]


async def test_xlsx_product_source_resolves_column_letters(tmp_path) -> None:
    path = tmp_path / "products.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(["mp", "sku_col", "url_col", "name_col"])
    sheet.append(["ozon", "999", "https://x/2", "Товар Б"])
    workbook.save(path)

    source = XlsxProductSource(
        products_path=str(path),
        products_sheet=None,
        products_range=None,
        regions_path=None,
        regions_sheet=None,
        regions_range=None,
        products_mapping={"marketplace": "A", "sku": "B", "url": "C", "name": "D"},
        regions_mapping=None,
    )

    rows = await source.read_products()

    assert rows == [{"marketplace": "ozon", "sku": "999", "url": "https://x/2", "name": "Товар Б"}]


async def test_xlsx_result_sink_writes_mapped_columns_with_decimal_price_as_string(tmp_path) -> None:
    out_path = tmp_path / "results.xlsx"
    sink = XlsxResultSink(
        path=str(out_path), sheet="Результаты", mapping={"sku": "A", "price": "B", "status": "C"}
    )

    written = await sink.write_snapshots([{"sku": "12345", "price": Decimal("199.90"), "status": "ok"}])

    assert written == 1
    workbook = openpyxl.load_workbook(out_path)
    sheet = workbook["Результаты"]
    assert [cell.value for cell in sheet[1]] == ["A", "B", "C"]
    assert [cell.value for cell in sheet[2]] == ["12345", "199.90", "ok"]
